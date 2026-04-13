#!/usr/bin/env python3
"""
Reads the Excel workbook and writes tools/excel-accessories-import/products.json + images/.

Usage:
  python3 wp-content/themes/mygun/tools/build-excel-import.py "/path/to/file.xlsx"
"""

from __future__ import annotations

import json
import os
import re
import sys
import zipfile

import openpyxl


def describe_ka(title: str) -> str:
    """Georgian product description from title (rule-based, unique per intent)."""
    t = title.strip()
    tl = t.lower()

    def p(*parts: str) -> str:
        return "<p>" + " ".join(parts) + "</p>"

    if "ყურის დამცავი" in t or "ear" in tl:
        return p(
            f"«{t}» სმენის დაცვის საშუალებაა — ამცირებს ხმაურსა და დისკომფორტს სროლის მოედანზე ან სპორტულ ვარჯიშზე.",
            "მსუბუქი, პრაქტიკული დიზაინი; შეარჩიეთ ფერი თქვენი აღჭურვილობის მიხედვით.",
        )
    if "სანათი" in t and "ოთხ" not in t and "მზის" not in t:
        return p(
            f"«{t}» კომპაქტური განათებაა ჩანთაში, მანქანაში ან სამუშაო სივრცეში.",
            "შესაფერისია სიბნელეში ორიენტაციისა და დეტალების დასანახად.",
        )
    if "მზის" in t or "განათება" in t:
        return p(
            f"«{t}» ენერგოეფექტული განათებაა მზის ელემენტით — კარგია კემპინგისა და გარე სამუშაოებისთვის.",
        )
    if "ოთხიანი" in t:
        return p(
            f"«{t}» უზრუნველყოფს ფართო განათებას; მზის პანელი ამცირებს ელექტროენერგიის მოხმარებას.",
        )
    if "წებოვანი" in t or "მიზანი" in t:
        return p(
            f"«{t}» სამიზნე ფირფიტების ან სხვა ზედაპირების სწრაფი მიმაგრებისთვისაა — ხშირად გამოიყენება სასწავლო და სპორტულ სროლაში.",
        )
    if "ჯეკი" in t or "jack" in tl:
        return p(
            f"«{t}» პორტატული აწევის ინსტრუმენტია — სასარგებლოა მანქანის საბურავის შეცვლისა და სასწრაფო სიტუაციებისთვის.",
        )
    if "შლანგი" in t:
        return p(
            f"«{t}» სასარგებლოა ჰაერის შევსებისა და ტექნიკური სამუშაოებისთვის; კომპლექტში შედის საჭირო თავები.",
        )
    if "პარალონი" in t:
        return p(
            f"«{t}» მრავალჯერადი დაცვითი ფენაა სამუშაო ადგილისთვის — იცავს ზედაპირს ნაკაწრისა და ჭუჭყისგან.",
        )
    if "ატვიორკ" in t or "emtop" in tl:
        return p(
            f"«{t}» ბოლტებისა და ქანჩების ნაკრებია — ერთ კომპლექტში გაქვთ სხვადასხვა ზომა ყოველდღიური შეკრებისა და შეკეთებისთვის.",
        )
    if "სკამი" in t:
        return p(
            f"«{t}» მობილური საჯდომია გარეთ სამუშაოებისა და დასვენებისთვის; მარტივად იკეცება და იტანს სიმძიმეს.",
        )
    if "ხერხი" in t:
        return p(
            f"«{t}» ხელსაწყოა ხის, პლასტმასის ან მსუბუქი მასალების დასაჭრელად — კომპაქტური ვერსია ჩანთაში გასატანად.",
        )
    if "სამიზნე" in t:
        return p(
            f"«{t}» სასწავლო სამიზნეა სროლის უნარების გასაუმჯობესებლად; დაიცავით უსაფრთხოების წესები და გამოიყენეთ დასაშვებ ტერიტორიაზე.",
        )
    if "დამჭერი" in t or "ჩამოსაკიდი" in t or "საკიდი" in t or "კედელზე" in t:
        return p(
            f"«{t}» ორგანიზებას უწყობს ხელს — იარაღის ან აღჭურვილობის უსაფრთხოდ შენახვისთვისაა განკუთვნილი.",
        )
    if "გილზების ჩასაყრელი" in t or "გილზების" in t:
        return p(
            f"«{t}» პრაქტიკული აქსესუარია ჭურჭლის მოწესრიგებისთვის; მაგრდება იარაღზე მითითებული სისტემით.",
        )
    if "ბალაკლავა" in t or "მაისური" in t or "ჟილეტი" in t:
        return p(
            f"«{t}» კომფორტული ტანსაცმელია გარე აქტივობისთვის — შეარჩიეთ ზომა ცხოვრების სტილის მიხედვით.",
        )
    if "ხელთათმანი" in t:
        return p(
            f"«{t}» იცავს ხელებს სიცივისა და ტრავმისგან; შესაფერისია სამუშაოსა და სპორტისთვის.",
        )
    if "ჩანთა" in t or "bag" in tl or "backpack" in tl or "ზურგჩანთა" in t:
        return p(
            f"«{t}» ტევადი ჩანთაა აღჭურვილობის გადასატანად — გაითვალისწინეთ განზომილება და განლაგება თქვენი ამოცანის მიხედვით.",
        )
    if "გონგი" in t:
        return p(
            f"«{t}» სასწავლო/სპორტული სამიზნეა სიზუსტის ვარჯიშისთვის — გამოიყენეთ უსაფრთხო დისტანციით.",
        )
    if "კლუჩ" in t or "ქანჩ" in t or "ნაკრები" in t or "ნაბორი" in t:
        return p(
            f"«{t}» მრავალფუნქციური ნაკრებია შეკრებისა და შეკეთებისთვის — ერთი კომპლექტი რამდენიმე ზომას ფარავს.",
        )
    if "დანის ჩასადები" in t or "ჩასადები" in t or "კაბურა" in t:
        return p(
            f"«{t}» უზრუნველყოფს უსაფრთხო ტარებასა და სწრაფ წვდომას — შეამოწმეთ თავსებადობა თქვენს მოდელთან.",
        )
    if "აკუმულატორი" in t or "დატენვ" in t:
        return p(
            f"«{t}» ენერგიის მობილური წყაროა მცირე მოწყობილობებისთვის — გამოიყენეთ მწარმოებლის ინსტრუქციის მიხედვით.",
        )
    if "საწმენდი" in t:
        return p(
            f"«{t}» ლულისა და მექანიზმის მოვლისთვისაა — რეგულარული წმენდა ახანგრძლივებს სერვისის ვადას.",
        )
    if "ჩახმახ" in t or "ბლოკი" in t:
        return p(
            f"«{t}» სათადარიგო ნაწილია მექანიზმის საიმედო მუშაობისთვის — დაამონტაჟეთ ხარისხიანი სპეციალისტის ან ინსტრუქციის მიხედვით.",
        )
    if "ტყვი" in t:
        return p(
            f"«{t}» ტევადი შენახვის აქსესუარია — ინახავს ჭურჭელს მოწესრილებულად და იცავს გარემოს ზედმეტი ხმაურისგან.",
        )
    if "ბიპოდ" in tl or "bipod" in tl:
        return p(
            f"«{t}» სტაბილურობას უზრუნველყოფს სროლისას — მარტივად ირეგულირება სიმაღლეზე და ზედაპირზე.",
        )
    if "red dot" in tl or "ლაზერ" in t:
        return p(
            f"«{t}» სამიზნის სისტემაა სწრაფი მორგებით — შეამოწმეთ კანონმდებლობა და თავსებადობა თქვენს იარაღთან.",
        )
    if "კამერა" in t or "1080" in t:
        return p(
            f"«{t}» უკანა ხედვის კამერაა მანქანისთვის — უზრუნველყოფს უფრო უსაფრთხო მანევრირებასა და პარკირებას.",
        )
    if "პიკატინი" in t or "თამას" in t:
        return p(
            f"«{t}» ალუმინის აქსესუარია რეილზე მონტაჟისთვის — მსუბუქი და გამძლე კონსტრუქცია.",
        )
    if "პერემიჩკა" in t or "2000a" in tl:
        return p(
            f"«{t}» საბატარეო გადამცემია განსაკუთრებული დატვირთვისთვის — მიადგით უსაფრთხოების წესებს.",
        )
    if "ელექტრო ქანჩი" in t:
        return p(
            f"«{t}» ელექტრო ინსტრუმენტია ქანჩების სწრაფი მონტაჟისთვის — შეამცირებს ფიზიკურ დატვირთვას.",
        )
    if "საზომი" in t and "ლაზერ" in t:
        return p(
            f"«{t}» მანძილის სწრაფი შეფასებისთვისაა — გამოსადეგია სპორტულ და ტექნიკურ ამოცანებში.",
        )
    if "საზომი" in t or "ლაზერული" in t:
        return p(
            f"«{t}» ზუსტი გაზომვის ინსტრუმენტია — მარტივი მართვა და კითხვადი ინდიკაცია.",
        )
    if "რეზინი" in t or "დასარბილებ" in t:
        return p(
            f"«{t}» ამცირებს ვიბრაციას და უზრუნველყოფს უფრო კომფორტულ ჭიდს — შეარჩიეთ ფერი თქვენი აღჭურვილობისთვის.",
        )
    if "ნიჩაბი" in t:
        return p(
            f"«{t}» პატარა ხელსაწყოა სიზუსტის სამუშაოებისთვის — კომპაქტული ზომა ჩანთაში გასატანად.",
        )
    if "სეკატორი" in t:
        return p(
            f"«{t}» უნივერსალური სეკატორია ბაღისა და ტექნიკური სამუშაოებისთვის.",
        )
    if "emergency" in tl:
        return p(
            f"«{t}» სასწრაფო ნაკრებია ძირითადი საჭირო ნივთებით — შეინახეთ მანქანაში ან სახლში სწრაფი წვდომისთვის.",
        )
    if "ნოუთბუქ" in t or "ტელეფონ" in t:
        return p(
            f"«{t}» ელექტრონიკის მოვლის ნაკრებია — სასარგებლოა სახლში და გზაზე.",
        )
    if "პლოცკო" in t:
        return p(
            f"«{t}» მობილური საჯდომი/მაგიდაა — იკეცება და არ იკავებს ბევრ ადგილს.",
        )
    if "ყურსასმენ" in t:
        return p(
            f"«{t}» იცავს ყურსასმენს დაზიანებისა და ჭუჭყისგან; პრაქტიკული ტარების ფორმა.",
        )
    if "m249" in tl:
        return p(
            f"«{t}» სპეციალიზებული ჩანთაა დიდი იარაღის გადასატანად — შეამოწმეთ ზომა და შიდა განყოფილებები.",
        )
    if "სნაიპერ" in t or "შაშხან" in t:
        return p(
            f"«{t}» გრძელი ჩანთაა სნაიპერული სისტემის ტრანსპორტირებისთვის — გაითვალისწინეთ სიგრძე და ბადეები.",
        )
    if "ვინჩესტერ" in t or "ეარსოფთ" in t:
        return p(
            f"«{t}» რბილი ჩანთაა ხელჯოხიანი იარაღის დასაცავად გადაადგილებისას.",
        )
    if "ქვიშის ჩანთა" in t:
        return p(
            f"«{t}» სროლის ვარჯიშისთვისაა — ეხმარება სტაბილურ დასადგამად და სამიზნის დაცვას.",
        )
    if "red dot" in tl:
        return p(
            f"«{t}» კოლიმატორული მიზანი ლაზერით — სწრაფი მიზანში მოყვანა სხვადასხვა განათებაში.",
        )

    return p(
        f"«{t}» აქსესუარების კატეგორიის ნივთია — ხარისხიანი შერჩევა ყოველდღიური გამოყენებისა და სპეციალიზებული ამოცანებისთვის.",
        "დეტალებისთვის დაგვიკავშირდით ან ეწვიეთ მაღაზიას.",
    )


def row_to_media_map(xlsx: str) -> dict[int, str]:
    with zipfile.ZipFile(xlsx) as z:
        rels = z.read("xl/drawings/_rels/drawing1.xml.rels").decode()
        rid_to_media: dict[str, str] = {}
        for m in re.finditer(r'Id="(rId\d+)"[^>]+Target="\.\./media/([^"]+)"', rels):
            rid_to_media[m.group(1)] = "xl/media/" + m.group(2)
        dxml = z.read("xl/drawings/drawing1.xml").decode()
    row_to_media: dict[int, str] = {}
    for block in re.finditer(r"<xdr:twoCellAnchor[^>]*>.*?</xdr:twoCellAnchor>", dxml, re.DOTALL):
        b = block.group(0)
        fm = re.search(r"<xdr:from>.*?<xdr:row>(\d+)</xdr:row>", b, re.DOTALL)
        em = re.search(r'r:embed="(rId\d+)"', b)
        if fm and em:
            rid = em.group(1)
            if rid in rid_to_media:
                row_to_media[int(fm.group(1)) + 1] = rid_to_media[rid]
    return row_to_media


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: build-excel-import.py /path/to/file.xlsx", file=sys.stderr)
        return 1
    xlsx = os.path.abspath(sys.argv[1])
    theme_tools = os.path.dirname(os.path.abspath(__file__))
    out_dir = os.path.join(theme_tools, "excel-accessories-import")
    img_dir = os.path.join(out_dir, "images")
    os.makedirs(img_dir, exist_ok=True)

    row_to_media = row_to_media_map(xlsx)
    products: list[dict] = []

    with zipfile.ZipFile(xlsx) as z:
        wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        ws = wb.active
        for r in range(2, ws.max_row + 1):
            title = ws.cell(r, 2).value
            if not title or not str(title).strip():
                continue
            title = str(title).strip()
            sale = ws.cell(r, 5).value
            try:
                price = float(sale) if sale is not None else 0.0
            except (TypeError, ValueError):
                price = 0.0
            media_path = row_to_media.get(r)
            img_name = None
            if media_path:
                ext = os.path.splitext(media_path)[1] or ".png"
                safe = f"row-{r}{ext}"
                with open(os.path.join(img_dir, safe), "wb") as fh:
                    fh.write(z.read(media_path))
                img_name = safe
            products.append(
                {
                    "excel_row": r,
                    "title": title,
                    "price": price,
                    "image": img_name,
                    "description": describe_ka(title),
                }
            )
        wb.close()

    payload = {
        "products": products,
        "category": {"slug": "aksesuarebi", "name": "აქსესუარები"},
    }
    with open(os.path.join(out_dir, "products.json"), "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    no_img = [p["excel_row"] for p in products if not p["image"]]
    print(f"Wrote {len(products)} products to {out_dir}; no image rows: {no_img}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
